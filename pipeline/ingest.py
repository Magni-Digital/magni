#!/usr/bin/env python3
"""
ingest.py — read practices from the two places Almendra puts them:

  1. inbox.csv            — new finds she drops in each day (one row per practice
                            she found on LinkedIn). Only company + website are
                            required; contact/email/vertical are optional.
  2. Magni_Digital_CRM.xlsx (the "Pipeline" sheet) — her existing CRM. Rows that
                            are already contacted are NOT re-qualified; they're
                            returned as prior "sent" dispositions so dedupe
                            retires them and they never resurface.

Output: (candidates, prior_dispositions)
  candidates           — list of normalized practice dicts to qualify today
  prior_dispositions   — [{key, status:'sent'}] for already-worked CRM rows
"""
from __future__ import annotations

import csv
from pathlib import Path

from .normalize import clean_name, domain_from_url, dedupe_key

# free-text vertical → canonical practice_type. Types in signals.HEALTH_TYPES
# get the "no online booking" check; everything else uses the universal signals.
_TYPE_MAP = [
    (("dental", "dentist", "ortho"), "dental"),
    (("therapy", "counsel", "psych", "mental", "behavioral"), "mental_health"),
    (("addiction", "recovery", "rehab center", "substance"), "addiction_recovery"),
    (("physical therapy", "physio", "rehabilitation", "rehab"), "physical_therapy"),
    (("chiro",), "chiropractic"),
    (("derm",), "dermatology"),
    (("med spa", "medspa", "aesthet", "estétic"), "med_spa"),
    (("wellness", "integrative", "naturopath", "acupunct", "bienestar"), "integrative_health"),
    (("optom", "eye care", "vision"), "optometry"),
    (("clinic", "medical", "specialty", "consultorio", "clínica"), "medical_clinic"),
    (("law", "legal", "attorney", "abogad"), "law"),
    (("financial", "wealth", "advisory", "account", "tax", "fiscal", "patrimonial"), "financial"),
    (("insurance", "seguros", "broker"), "insurance"),
    (("coach",), "coaching"),
    (("consult",), "consulting"),
    (("staffing", "recruit", "hr ", "human resources"), "staffing"),
    (("school", "tutor", "education", "colegio", "idiomas"), "education"),
    (("real estate", "realty", "brokerage", "inmobil"), "real_estate"),
    (("nonprofit", "non-profit", "foundation"), "nonprofit"),
]

# CRM Status values that mean "already worked" → don't re-surface.
_CONTACTED_STATUSES = {"1st sent", "follow-up sent", "replied", "in conversation",
                       "call booked", "won", "lost", "do not contact"}


def classify(*texts):
    blob = " ".join(t for t in texts if t).lower()
    for needles, canon in _TYPE_MAP:
        if any(n in blob for n in needles):
            return canon
    return ""   # unknown / generic — universal signals still apply


def _candidate(name, *, website="", contact="", role="", email="", source="",
               location="", vertical="", notes="", prior_observation=""):
    name = (name or "").strip()
    domain = domain_from_url(website)
    return {
        "name": name,
        "contact_name": clean_name(contact),
        "role": (role or "").strip(),
        "website_raw": (website or "").strip(),
        "domain": domain,
        "email": (email or "").strip(),
        "source": (source or "").strip(),
        "location": (location or "").strip(),
        "practice_type": classify(vertical, name, notes),
        "notes": (notes or "").strip(),
        "prior_observation": (prior_observation or "").strip(),
        "website_status": "none" if not (website or "").strip() else "unknown",
    }


def read_inbox(path):
    p = Path(path)
    if not p.exists():
        return []
    out = []
    with open(p, encoding="utf-8-sig", errors="replace") as fh:
        for row in csv.DictReader(fh):
            row = {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
            name = row.get("company") or row.get("name") or ""
            if not name and not row.get("website"):
                continue
            out.append(_candidate(
                name, website=row.get("website", ""),
                contact=row.get("contact_name") or row.get("contact", ""),
                role=row.get("role", ""), email=row.get("email", ""),
                source=row.get("source", "inbox"),
                location=row.get("location", ""), vertical=row.get("vertical", ""),
                notes=row.get("notes", "")))
    return out


def read_crm(path, sheet="Pipeline"):
    """Return (fresh_candidates, prior_dispositions). Uncontacted rows become
    candidates; contacted rows become 'sent' dispositions for the seen-set."""
    p = Path(path)
    if not p.exists():
        return [], []
    import openpyxl
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return [], []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], []
    header = [(str(h).strip() if h is not None else "") for h in rows[0]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    ix = {
        "name": col("Company", "Company Name"),
        "contact": col("Contact Name", "Contact"),
        "role": col("Role", "Title"),
        "website": col("Website URL", "Website"),
        "obs": col("Site Observation (1 line)", "Site Observation"),
        "email": col("Email"),
        "source": col("Source"),
        "status": col("Status"),
        "notes": col("Notes"),
    }

    def cell(r, key):
        i = ix.get(key)
        return ("" if i is None or i >= len(r) or r[i] is None else str(r[i]).strip())

    fresh, prior = [], []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        name = cell(r, "name")
        website = cell(r, "website")
        if not name and not website:
            continue
        cand = _candidate(
            name, website=website, contact=cell(r, "contact"), role=cell(r, "role"),
            email=cell(r, "email"), source=cell(r, "source") or "CRM",
            notes=cell(r, "notes"), prior_observation=cell(r, "obs"))
        status = cell(r, "status").lower()
        if status in _CONTACTED_STATUSES:
            key = dedupe_key(cand["domain"], name=cand["name"], location=cand["location"])
            if key:
                prior.append({"key": key, "status": "sent"})
        else:
            fresh.append(cand)
    return fresh, prior


def _identity(c):
    return dedupe_key(c["domain"], name=c["name"], location=c["location"])


def load_candidates(inbox_path, crm_path):
    """Merge inbox + CRM, de-duplicating within this run by identity key.
    Returns (candidates, prior_dispositions)."""
    inbox = read_inbox(inbox_path)
    crm_fresh, prior = read_crm(crm_path)
    merged, seen_keys = [], set()
    for c in inbox + crm_fresh:        # inbox wins on conflict (freshest source)
        k = _identity(c)
        if not k or k in seen_keys:
            continue
        seen_keys.add(k)
        merged.append(c)
    return merged, prior
